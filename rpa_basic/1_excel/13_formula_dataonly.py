from openpyxl import load_workbook
# wb = load_workbook("sample_formula.xlsx")
# ws = wb.active

# 수식 그대로 가져옴
# for row in ws.values:
    # for cell in row:
        # print(cell)
        
wb = load_workbook("sample_formula.xlsx", data_only=True)
ws = wb.active

# 수식 아닌 실제 데이터 가지고 옴
# evaluate 되지 않은 상태의 데이터는 None이라고 표시
for row in ws.values:
    for cell in row:
        print(cell)
