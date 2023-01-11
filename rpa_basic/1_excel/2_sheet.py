from openpyxl import Workbook
wb = Workbook()
ws = wb.create_sheet()  # 새로운sheet 기본이름으로 생성
ws.title = "MySheet"  # Sheet 이름 변경
ws.sheet_properties.tabColor = "cc33ff"  # rgb형태로 값을 넣어주면 탭 색상 변경

# sheet, mysheet, yoursheet
ws1 = wb.create_sheet("YourSheet") # 주어진 이름으로 sheet생성
ws2 = wb.create_sheet("NewSheet", 2) # 2번째 index에 sheet 생성

new_ws = wb["NewSheet"] # dic형태로 sheet에 접근

print(wb.sheetnames) # 모든 sheet이름 확인

# sheet 복사
new_ws["A1"] = "Test"
target = wb.copy_worksheet(new_ws)
target.title = "Copied Sheet"


wb.save("sample.xlsx")
