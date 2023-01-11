from openpyxl import load_workbook
wb = load_workbook("sample.xlsx")
ws = wb.active

# 번호 영어 수학
# 번호 (국어) 영어 수학
# ws.move_range("B1:C11", rows=0,cols=1) # 열기준 오른쪽으로 1칸 이동
# ws["B1"].value = "국어" # b1셀에 '국어' 입력

# 번호 영어 수학
ws.move_range("C1:C11", rows=5, cols=-1) # 5줄 내려가고 왼쪽으로 1칸 이동: 내용이 덮어씌워짐

wb.save("sample_korean.xlsx")
