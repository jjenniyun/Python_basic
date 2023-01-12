# 퀴즈 과목의 점수 비중
# 출석 10 퀴즈10 퀴즈2 10 중간고사 20 기말고사 30 프로젝트 20 총합 100
# 퀴즈2 오류 발견하여 모두 만점 처리 하기로 함 # 현재까지 작성된 성적 데이터 수정하시오
# 조건 퀴즈 2 점수 10 수정
# H열에 총점(sum이용) . I열에 성적정보 추가
# - 총점 90점 이상 A, 80이상 B, 70점 이상 C, 나머지 D
# 출석이 5미만 학생은 총점 상관없이 F
# 최종 파일명 scores.xlsx

# [현재까지 작성된 최종 성적 데이터]
# 학번, 출석, 퀴즈1, 퀴즈2, 중간고사, 기말고사, 프로젝트
# 1,10,8,5,14,26,12
# 2,7,3,7,15,24,18
# 3,9,5,8,8,12,4
# 4,7,8,7,17,21,18
# 5,7,8,7,16,25,15
# 6,3,5,8,8,17,0
# 7,4,9,10,16,27,18
# 8,6,6,6,15,19,17
# 9,10,10,9,19,30,19
# 10,9,8,8,20,25,20

from openpyxl import Workbook
wb = Workbook()
ws = wb.active

# 현재까지 작성된 최종 성적 데이터 넣기
ws.append(("학번", "출석", "퀴즈1", "퀴즈2", "중간고사", "기말고사", "프로젝트"))

scores = [
(1,10,8,5,14,26,12),
(2,7,3,7,15,24,18),
(3,9,5,8,8,12,4),
(4,7,8,7,17,21,18),
(5,7,8,7,16,25,15),
(6,3,5,8,8,17,0),
(7,4,9,10,16,27,18),
(8,6,6,6,15,19,17),
(9,10,10,9,19,30,19),
(10,9,8,8,20,25,20)
]

for s in scores: # 기본 성적 데이터 넣기
    ws.append(s)
    
# 퀴즈 2 점수 10 수정
for idx, cell in enumerate(ws["D"]):
    if idx == 0: # 제목인 경우 skip
        continue
    cell.value = 10
    
# H열에 총점(sum이용) . I열에 성적정보 추가
ws["H1"] = "총점"
ws["I1"] = "성적"

# a b c d e f g h i
# 1 2 3 4 5 6 7 8 9
for idx, score in enumerate(scores, start=2):
    sum_val = sum(score[1:]) - score[3] + 10 # 퀴즈2 점수 제외하고 10점만 더해주기 : 총점
    ws.cell(row=idx, column=8).value = "=SUM(B{}:G{})".format(idx, idx)
    # SUM(B2:G2)
    # SUM(B3:G3) ...
    
    # - 총점 90점 이상 A, 80이상 B, 70점 이상 C, 나머지 D
    grade = None # 성적
    if sum_val >= 90:
        grade = "A"
    elif sum_val >= 80:
        grade = "B"
    elif sum_val >= 70:
        grade = "C"
    else:
        grade = "D"
        
    # 출석이 5미만 학생은 총점 상관없이 F
    if score[1] < 5:
        grade = "F"
        
    ws.cell(row=idx, column=9).value = grade # I열에 성적 정보
    
wb.save("scores.xlsx")